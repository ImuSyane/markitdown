"""
Enhanced XLSX Converter with OCR support for embedded images.
Extracts images from Excel spreadsheets and performs OCR while maintaining cell context.
"""

import io
import sys
from typing import Any, BinaryIO, Optional

from markitdown.converters import HtmlConverter
from markitdown import DocumentConverter, DocumentConverterResult, StreamInfo
from markitdown._exceptions import (
    MissingDependencyException,
    MISSING_DEPENDENCY_MESSAGE,
)
from ._image_export import create_image_asset, image_export_enabled, render_image_block
from ._ocr_service import OCRService

# Try loading dependencies
_xlsx_dependency_exc_info = None
try:
    import pandas as pd
    from openpyxl import load_workbook
except ImportError:
    _xlsx_dependency_exc_info = sys.exc_info()


class XlsxConverterWithOCR(DocumentConverter):
    """
    Enhanced XLSX Converter with OCR support for embedded images.
    Extracts images with their cell positions and performs OCR.
    """

    def __init__(self, ocr_service: Optional[OCRService] = None):
        super().__init__()
        self._html_converter = HtmlConverter()
        self.ocr_service = ocr_service

    def accepts(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,
    ) -> bool:
        mimetype = (stream_info.mimetype or "").lower()
        extension = (stream_info.extension or "").lower()

        if extension == ".xlsx":
            return True

        if mimetype.startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml"
        ):
            return True

        return False

    def convert(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,
    ) -> DocumentConverterResult:
        if _xlsx_dependency_exc_info is not None:
            raise MissingDependencyException(
                MISSING_DEPENDENCY_MESSAGE.format(
                    converter=type(self).__name__,
                    extension=".xlsx",
                    feature="xlsx",
                )
            ) from _xlsx_dependency_exc_info[1].with_traceback(
                _xlsx_dependency_exc_info[2]
            )  # type: ignore[union-attr]

        # Get OCR service if available (from kwargs or instance)
        ocr_service: Optional[OCRService] = (
            kwargs.get("ocr_service") or self.ocr_service
        )

        if ocr_service or image_export_enabled(**kwargs):
            # Remove ocr_service from kwargs to avoid duplicate argument error
            kwargs_without_ocr = {k: v for k, v in kwargs.items() if k != "ocr_service"}
            return self._convert_with_images(
                file_stream, ocr_service, **kwargs_without_ocr
            )
        else:
            return self._convert_standard(file_stream, **kwargs)

    def _convert_standard(
        self, file_stream: BinaryIO, **kwargs: Any
    ) -> DocumentConverterResult:
        """Standard conversion without OCR."""
        file_stream.seek(0)
        sheets = pd.read_excel(file_stream, sheet_name=None, engine="openpyxl")
        md_content = ""

        for sheet_name in sheets:
            md_content += f"## {sheet_name}\n"
            html_content = sheets[sheet_name].to_html(index=False)
            md_content += (
                self._html_converter.convert_string(
                    html_content, **kwargs
                ).markdown.strip()
                + "\n\n"
            )

        return DocumentConverterResult(markdown=md_content.strip())

    def _convert_with_images(
        self, file_stream: BinaryIO, ocr_service: Optional[OCRService], **kwargs: Any
    ) -> DocumentConverterResult:
        """Convert XLSX with inline image export and optional OCR."""
        file_stream.seek(0)
        wb = load_workbook(file_stream)

        md_content = ""
        assets = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            md_content += f"## {sheet_name}\n\n"

            # Convert sheet data to markdown table
            file_stream.seek(0)
            try:
                df = pd.read_excel(
                    file_stream, sheet_name=sheet_name, engine="openpyxl"
                )
                html_content = df.to_html(index=False)
                md_content += (
                    self._html_converter.convert_string(
                        html_content, **kwargs
                    ).markdown.strip()
                    + "\n\n"
                )
            except Exception:
                # If pandas fails, just skip the table
                pass

            # Extract and OCR images in this sheet
            images_with_ocr = self._extract_and_ocr_sheet_images(
                sheet,
                sheet_name=sheet_name,
                ocr_service=ocr_service,
                image_dir=kwargs.get("image_dir"),
            )

            if images_with_ocr:
                md_content += "### Images in this sheet:\n\n"
                for img_info in images_with_ocr:
                    assets.extend(img_info["assets"])
                    if img_info["asset"] is not None:
                        md_content += (
                            render_image_block(
                                img_info["asset"],
                                alt_text=img_info["cell_ref"],
                                extracted_text=img_info["ocr_text"],
                            )
                            + "\n\n"
                        )
                    elif img_info["ocr_text"]:
                        md_content += (
                            f"*[Image OCR]\n{img_info['ocr_text']}\n[End OCR]*\n\n"
                        )

        return DocumentConverterResult(markdown=md_content.strip(), assets=assets)

    def _extract_and_ocr_sheet_images(
        self,
        sheet: Any,
        *,
        sheet_name: str,
        ocr_service: Optional[OCRService],
        image_dir: Optional[str],
    ) -> list[dict]:
        """
        Extract and OCR images from an Excel sheet.

        Args:
            sheet: openpyxl worksheet
            ocr_service: OCR service

        Returns:
             List of dicts with image metadata, optional OCR, and exported assets
        """
        results = []

        try:
            # Check if sheet has images
            if hasattr(sheet, "_images"):
                for index, img in enumerate(sheet._images, start=1):
                    try:
                        # Get image data
                        if hasattr(img, "_data"):
                            image_data = img._data()
                        elif hasattr(img, "image"):
                            # Some versions store it differently
                            image_data = img.image
                        else:
                            continue

                        # Create image stream
                        image_stream = io.BytesIO(image_data)

                        # Get cell reference
                        cell_ref = "unknown"
                        if hasattr(img, "anchor"):
                            anchor = img.anchor
                            if hasattr(anchor, "_from"):
                                from_cell = anchor._from
                                if hasattr(from_cell, "col") and hasattr(
                                    from_cell, "row"
                                ):
                                    # Convert column number to letter
                                    col_letter = self._column_number_to_letter(
                                        from_cell.col
                                    )
                                    cell_ref = f"{col_letter}{from_cell.row + 1}"

                        asset = None
                        assets = []
                        image_format = getattr(img, "format", None)
                        image_extension = (
                            f".{str(image_format).lower()}" if image_format else None
                        )
                        image_mimetype = (
                            f"image/{str(image_format).lower()}" if image_format else None
                        )
                        if image_dir:
                            asset = create_image_asset(
                                image_bytes=image_data,
                                image_dir=image_dir,
                                name=f"{sheet_name}_{cell_ref}_{index}",
                                extension=image_extension,
                                mimetype=image_mimetype,
                            )
                            assets.append(asset)

                        ocr_text = ""
                        if ocr_service:
                            ocr_result = ocr_service.extract_text(image_stream)
                            ocr_text = ocr_result.text.strip()

                        if asset is not None or ocr_text:
                            results.append(
                                {
                                    "cell_ref": cell_ref,
                                    "ocr_text": ocr_text,
                                    "asset": asset,
                                    "assets": assets,
                                }
                            )

                    except Exception:
                        continue

        except Exception:
            pass

        return results

    @staticmethod
    def _column_number_to_letter(n: int) -> str:
        """Convert column number to Excel column letter (0-indexed)."""
        result = ""
        n = n + 1  # Make 1-indexed
        while n > 0:
            n -= 1
            result = chr(65 + (n % 26)) + result
            n //= 26
        return result
